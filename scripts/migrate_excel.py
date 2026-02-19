#!/usr/bin/env python3
"""
Excel請求書データ → Cloudflare D1 移行スクリプト
Usage: python migrate_excel.py --file <excel_path> --output <sql_path>
"""
import argparse
import math
import re
import uuid
import logging
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# 通貨シンボルのマッピング
CURRENCY_SYMBOLS = {
    'A$': 'AUD', 'C$': 'CAD', 'S$': 'SGD', 'NZ$': 'NZD',
    '$': 'USD', '€': 'EUR', '£': 'GBP', '¥': 'JPY',
}

# 請求書キーワード（日英）
INVOICE_NUMBER_KEYS = ['invoice no', 'invoice #', '請求書番号', '請求番号', 'inv no', 'inv #']
DATE_KEYS = ['invoice date', '請求日', '発行日', 'date', '日付']
DUE_DATE_KEYS = ['due date', '支払期限', '支払い期限', 'payment due', 'due']
TOTAL_KEYS = ['total', '合計', '請求合計', '小計合計', 'grand total', '請求金額']
TAX_RATE_KEYS = ['tax', '税率', 'vat', 'gst', '消費税']


def detect_currency(value_str: str) -> tuple[str, float]:
    """文字列から通貨コードと金額（float）を抽出する。"""
    if value_str is None:
        return 'USD', 0.0
    s = str(value_str).strip().replace(',', '').replace(' ', '')
    # 長いシンボルを先にマッチ
    for symbol, code in CURRENCY_SYMBOLS.items():
        if s.startswith(symbol):
            num_part = s[len(symbol):]
            try:
                return code, float(num_part)
            except ValueError:
                return code, 0.0
    # シンボルなし → 数値のみ
    try:
        return 'USD', float(s)
    except ValueError:
        return 'USD', 0.0


def cell_str(cell) -> str:
    """セル値を文字列として返す。Noneは空文字。"""
    if cell is None or cell.value is None:
        return ''
    return str(cell.value).strip()


def find_keyword_cell(ws, keywords: list[str]) -> tuple[int, int] | None:
    """シート内でキーワードにマッチするセルの (row, col) を返す。"""
    for row in ws.iter_rows():
        for cell in row:
            val = cell_str(cell).lower()
            for kw in keywords:
                if kw in val:
                    return cell.row, cell.column
    return None


def extract_value_near(ws, keywords: list[str], search_rows: int = 30) -> str:
    """キーワードセルの右または下のセルから値を取得する。"""
    loc = find_keyword_cell(ws, keywords)
    if loc is None:
        return ''
    r, c = loc
    # 右のセルを確認
    right = ws.cell(row=r, column=c + 1)
    if right.value is not None:
        return cell_str(right)
    # 下のセルを確認
    below = ws.cell(row=r + 1, column=c)
    if below.value is not None:
        return cell_str(below)
    return ''


def parse_date(value_str: str) -> str | None:
    """様々な形式の日付文字列を ISO 8601 (YYYY-MM-DD) に変換する。"""
    if not value_str:
        return None
    # datetime オブジェクトが文字列化されている場合
    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y',
        '%d-%m-%Y', '%B %d, %Y', '%b %d, %Y',
        '%Y年%m月%d日', '%Y-%m-%d %H:%M:%S',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    # Excel シリアル日付の場合（openpyxl は通常 datetime に変換するが念のため）
    try:
        n = float(value_str)
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=n)).strftime('%Y-%m-%d')
    except (ValueError, OverflowError):
        pass
    return None


def detect_item_rows(ws) -> list[dict]:
    """
    明細行を検出する。
    「品名 / 単価 / 数量」または「Description / Unit Price / Quantity」のような
    ヘッダー行を探し、その下の連続するデータ行を取得する。
    """
    ITEM_HEADER_KEYWORDS = [
        'description', 'item', '品名', '内容', '項目',
        'qty', 'quantity', '数量',
        'unit price', 'unit cost', '単価',
        'amount', '金額',
    ]

    header_row = None
    col_map = {}  # キー: 'name' | 'qty' | 'unit_price' | 'amount' | 'tax_rate'

    for row in ws.iter_rows(max_row=50):
        matched = 0
        tmp_map = {}
        for cell in row:
            val = cell_str(cell).lower()
            if any(kw in val for kw in ['description', 'item', '品名', '内容', '項目']):
                tmp_map['name'] = cell.column
                matched += 1
            elif any(kw in val for kw in ['qty', 'quantity', '数量']):
                tmp_map['qty'] = cell.column
                matched += 1
            elif any(kw in val for kw in ['unit price', 'unit cost', '単価']):
                tmp_map['unit_price'] = cell.column
                matched += 1
            elif any(kw in val for kw in ['amount', '金額', '小計']):
                tmp_map['amount'] = cell.column
                matched += 1
            elif any(kw in val for kw in ['tax', '税率', 'vat', 'gst']):
                tmp_map['tax_rate'] = cell.column
                matched += 1
        if matched >= 2:
            header_row = row[0].row
            col_map = tmp_map
            break

    if header_row is None or 'name' not in col_map:
        return []

    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=col_map['name'])
        if not name_cell.value:
            # 空行 → 明細終了
            break

        name = cell_str(name_cell)
        # 合計行っぽい行はスキップ
        if any(kw in name.lower() for kw in ['total', '合計', 'subtotal', '小計']):
            continue

        qty_raw = cell_str(ws.cell(row=r, column=col_map.get('qty', 0))) if 'qty' in col_map else '1'
        unit_raw = cell_str(ws.cell(row=r, column=col_map.get('unit_price', 0))) if 'unit_price' in col_map else '0'
        tax_raw = cell_str(ws.cell(row=r, column=col_map.get('tax_rate', 0))) if 'tax_rate' in col_map else '0'

        try:
            qty = float(re.sub(r'[^\d.]', '', qty_raw) or '1')
        except ValueError:
            qty = 1.0

        _, unit_cost = detect_currency(unit_raw)
        if unit_cost == 0.0 and 'amount' in col_map:
            _, unit_cost = detect_currency(cell_str(ws.cell(row=r, column=col_map['amount'])))

        try:
            tax_str = re.sub(r'[^\d.]', '', tax_raw)
            tax_rate = float(tax_str) / 100.0 if tax_str else 0.0
            # 既に 0.1 形式の場合は変換不要
            if tax_rate > 1.0:
                tax_rate = tax_rate / 100.0
        except ValueError:
            tax_rate = 0.0

        # 端数処理: math.floor(unit_cost * qty * (1 + tax_rate))
        amount = math.floor(unit_cost * qty * (1 + tax_rate))

        items.append({
            'name': name,
            'quantity': qty,
            'unit_cost': unit_cost,
            'tax_rate': tax_rate,
            'amount': amount,
        })

    return items


def parse_invoice_sheet(ws) -> dict | None:
    """
    シートから請求書データを解析する。
    解析できない場合は None を返す。
    """
    # 最低限、日付または合計が必要
    invoice_number = extract_value_near(ws, INVOICE_NUMBER_KEYS)
    date_str = extract_value_near(ws, DATE_KEYS)
    due_date_str = extract_value_near(ws, DUE_DATE_KEYS)
    total_raw = extract_value_near(ws, TOTAL_KEYS)

    # セル値が datetime の場合も考慮
    date_loc = find_keyword_cell(ws, DATE_KEYS)
    if date_loc:
        right_val = ws.cell(row=date_loc[0], column=date_loc[1] + 1).value
        if isinstance(right_val, datetime):
            date_str = right_val.strftime('%Y-%m-%d')

    due_loc = find_keyword_cell(ws, DUE_DATE_KEYS)
    if due_loc:
        right_val = ws.cell(row=due_loc[0], column=due_loc[1] + 1).value
        if isinstance(right_val, datetime):
            due_date_str = right_val.strftime('%Y-%m-%d')

    issue_date = parse_date(date_str)
    due_date = parse_date(due_date_str)
    currency, total_amount = detect_currency(total_raw)

    # 明細行を解析
    items = detect_item_rows(ws)

    # 明細から合計を再計算
    if not items and total_amount == 0.0:
        logger.warning(f"  シート '{ws.title}': 有効な請求データが見つかりません。スキップします。")
        return None

    if items and total_amount == 0.0:
        total_amount = sum(i['amount'] for i in items)

    return {
        'invoice_number': invoice_number or ws.title,
        'issue_date': issue_date or datetime.now().strftime('%Y-%m-%d'),
        'due_date': due_date,
        'currency': currency,
        'total_amount': int(total_amount),
        'status': 'paid',  # 既存データは支払済みとみなす
        'notes': f'Migrated from sheet: {ws.title}',
        'items': items,
    }


def slugify_client_name(sheet_name: str) -> str:
    """シート名からクライアント名を推定する。"""
    # 日付パターン（例: 20240101, 2024-01, Jan2024）を除去
    name = re.sub(r'\d{4}[-/]?\d{0,2}[-/]?\d{0,2}', '', sheet_name)
    name = re.sub(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s*\d*', '', name, flags=re.IGNORECASE)
    name = name.strip(' -_/')
    return name if name else sheet_name


def escape_sql(value: str) -> str:
    """SQL文字列のシングルクォートをエスケープする。"""
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''") + "'"


def generate_sql(clients: list[dict], invoices: list[dict], items: list[dict]) -> str:
    """Cloudflare D1 互換の SQLite INSERT 文を生成する。"""
    lines = ['BEGIN TRANSACTION;', '']

    lines.append('-- =====================')
    lines.append('-- Clients')
    lines.append('-- =====================')
    for c in clients:
        lines.append(
            f"INSERT OR IGNORE INTO clients (id, name, created_at, updated_at) VALUES ("
            f"{escape_sql(c['id'])}, {escape_sql(c['name'])}, "
            f"{escape_sql(c['created_at'])}, {escape_sql(c['updated_at'])});"
        )

    lines.append('')
    lines.append('-- =====================')
    lines.append('-- Invoices')
    lines.append('-- =====================')
    for inv in invoices:
        due = escape_sql(inv['due_date']) if inv['due_date'] else 'NULL'
        lines.append(
            f"INSERT OR IGNORE INTO invoices "
            f"(id, client_id, invoice_number, issue_date, due_date, "
            f"currency, total_amount, status, notes, created_at, updated_at) VALUES ("
            f"{escape_sql(inv['id'])}, {escape_sql(inv['client_id'])}, "
            f"{escape_sql(inv['invoice_number'])}, {escape_sql(inv['issue_date'])}, "
            f"{due}, {escape_sql(inv['currency'])}, {inv['total_amount']}, "
            f"{escape_sql(inv['status'])}, {escape_sql(inv['notes'])}, "
            f"{escape_sql(inv['created_at'])}, {escape_sql(inv['updated_at'])});"
        )

    lines.append('')
    lines.append('-- =====================')
    lines.append('-- Invoice Items')
    lines.append('-- =====================')
    for item in items:
        lines.append(
            f"INSERT OR IGNORE INTO invoice_items "
            f"(id, invoice_id, name, quantity, unit_cost, tax_rate, amount, created_at) VALUES ("
            f"{escape_sql(item['id'])}, {escape_sql(item['invoice_id'])}, "
            f"{escape_sql(item['name'])}, {item['quantity']}, {item['unit_cost']}, "
            f"{item['tax_rate']}, {item['amount']}, {escape_sql(item['created_at'])});"
        )

    lines.append('')
    lines.append('COMMIT;')
    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Excel請求書 → Cloudflare D1 移行ツール')
    parser.add_argument('--file', required=True, help='Excelファイルパス (.xlsx)')
    parser.add_argument('--output', default='migration.sql', help='出力SQLファイルパス（デフォルト: migration.sql）')
    parser.add_argument('--dry-run', action='store_true', help='SQLを表示するだけで出力ファイルを作成しない')
    parser.add_argument('--default-currency', default='USD', help='デフォルト通貨コード（デフォルト: USD）')
    args = parser.parse_args()

    excel_path = Path(args.file)
    if not excel_path.exists():
        logger.error(f"ファイルが見つかりません: {excel_path}")
        exit(1)

    logger.info(f"Excelファイルを読み込み中: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        logger.error(f"Excelファイルの読み込みに失敗しました: {e}")
        exit(1)

    now_iso = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
    clients_map: dict[str, dict] = {}  # client_name -> client dict
    invoices: list[dict] = []
    all_items: list[dict] = []

    total_sheets = len(wb.sheetnames)
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"解析中: '{sheet_name}'")

        invoice_data = parse_invoice_sheet(ws)
        if invoice_data is None:
            skipped += 1
            continue

        # クライアント名を推定・登録
        client_name = slugify_client_name(sheet_name) or sheet_name
        if client_name not in clients_map:
            clients_map[client_name] = {
                'id': str(uuid.uuid4()),
                'name': client_name,
                'created_at': now_iso,
                'updated_at': now_iso,
            }
        client = clients_map[client_name]

        # 請求書レコード
        invoice_id = str(uuid.uuid4())
        invoices.append({
            'id': invoice_id,
            'client_id': client['id'],
            'invoice_number': invoice_data['invoice_number'],
            'issue_date': invoice_data['issue_date'],
            'due_date': invoice_data['due_date'],
            'currency': invoice_data['currency'],
            'total_amount': invoice_data['total_amount'],
            'status': invoice_data['status'],
            'notes': invoice_data['notes'],
            'created_at': now_iso,
            'updated_at': now_iso,
        })

        # 明細レコード
        for item in invoice_data['items']:
            all_items.append({
                'id': str(uuid.uuid4()),
                'invoice_id': invoice_id,
                'name': item['name'],
                'quantity': item['quantity'],
                'unit_cost': item['unit_cost'],
                'tax_rate': item['tax_rate'],
                'amount': item['amount'],
                'created_at': now_iso,
            })

    clients_list = list(clients_map.values())
    sql = generate_sql(clients_list, invoices, all_items)

    # サマリー表示
    logger.info(f"\n=== 移行サマリー ===")
    logger.info(f"総シート数    : {total_sheets}")
    logger.info(f"スキップ      : {skipped}")
    logger.info(f"クライアント数: {len(clients_list)}")
    logger.info(f"請求書数      : {len(invoices)}")
    logger.info(f"明細行数      : {len(all_items)}")

    if args.dry_run:
        print('\n' + sql)
        logger.info("\n--dry-run モード: ファイルへの書き込みをスキップしました。")
    else:
        output_path = Path(args.output)
        output_path.write_text(sql, encoding='utf-8')
        logger.info(f"\nSQLファイルを出力しました: {output_path.resolve()}")
        logger.info(f"次のステップ: README.md の「D1への投入方法」を参照してください。")


if __name__ == '__main__':
    main()
